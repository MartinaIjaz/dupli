"""
Генерира празна Excel табела (template) со сите полиња од прашалникот.
Фајлот е подготвен за рачно пополнување.

Употреба:
    python3 create_blank_template.py [--output template.xlsx] [--rows 50]
"""

import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Колони и нивните dropdown опции
# ---------------------------------------------------------------------------

COLUMNS = [
    "REDI Staff Member Name",
    "Date of Mapping",
    "Name Surname",
    "Contact Email",
    "Contact Phone Number",
    "Country",
    "Place of Residence (Municipality / City)",
    "This will help us to understand where our clients are living – Address",
    "Gender",
    "Year of Birth",
    "Level of Education",
    "Do you identify as Roma?",
    "If NO Roma – Connection to Roma Community",
    "Business Name",
    "Address of the Business",
    "Short Description of the Business",
    "Business Sector",
    "Is the Business Registered?",
    "Date of Business Registration",
    "Number of Employees (formal and informal aggregate)",
    "Was Business Supported in Phase I?",
    "Year of Initial Support",
    "Type of Service Needed",
    "Interested in Business Club Membership?",
    "Previously had a Business Loan?",
    "Informed about Measures / Help for Business",
    "[REDI Staff] Assessment of Client Needs",
    "[REDI Staff] Action Recommended",
]

DROPDOWNS = {
    "Level of Education": [
        "No formal education",
        "Elementary School",
        "High School",
        "University Student",
        "Bachelor",
        "Master",
        "PhD",
        "Vocational Training",
    ],
    "Do you identify as Roma?": ["Yes", "No"],
    "Business Sector": [
        "Agriculture and animal husbandry",
        "Construction",
        "Services (Accounting, Law)",
        "Tourism and Hospitality",
        "Beauty and Health",
        "Transportation",
        "Production",
        "Trade",
        "Craftsmanship",
        "Recycling & Upcycling",
        "Cleaning Agency",
        "Other",
    ],
    "Is the Business Registered?": ["Yes", "No"],
    "Was Business Supported in Phase I?": ["Yes", "No"],
    "Year of Initial Support": [
        "2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027",
    ],
    "Type of Service Needed": [
        "Incubation Program",
        "Acceleration Program",
        "Growth Program",
        "Digitalization of the business",
        "One-on-one Mentorship",
        "Business Registration",
        "Brand Identity",
        "Business Plan Writing",
        "Marketing and Promotion",
        "Networking",
        "Access to Grant from National Employment Service Agency",
        "Access to Grant from REDI",
        "Access to grant from IPARD",
        "Business Loan",
        "Certified Training",
        "Consulting",
        "Legal Support",
        "Accountancy Support",
        "Assistance at Finding Additional Workers",
        "Assistance at Starting a Business",
        "Internship Program Development",
        "Improved Business Practices",
        "Assistance in Accessing Government Institutions",
        "Assistance in Forming Cooperatives",
        "Social Impact and Corporate Social Responsibility (ESG)",
        "Nothing from above",
    ],
    "Interested in Business Club Membership?": ["Yes", "No", "Maybe"],
    "Previously had a Business Loan?": ["Yes", "No"],
    "Informed about Measures / Help for Business": ["Yes", "No"],
    "[REDI Staff] Action Recommended": [
        "Develop a business plan",
        "Link to a financial institution (MFI or Bank)",
        "Apply for REDI Grant Scheme Support",
        "Link to the Agency of Employment",
        "Link to Government measures, to be self-employed",
        "Train on Business Growth and Development",
        "Business Formalization (Registration)",
        "One-on-One Expert Support",
        "Digitalization (Roma Digital Boost)",
        "Marketing and Promotion",
        "Networking (Business Clubs)",
        "Linking to potential employees",
        "Business Consultancy",
        "Link clients to training programs",
        "No Action for the moment",
        "Other",
    ],
}

# ---------------------------------------------------------------------------
# Групи на колони → бои
# ---------------------------------------------------------------------------

def _col_group(col: str) -> str:
    if col.startswith("[REDI"):
        return "redi"
    if col in ("REDI Staff Member Name", "Date of Mapping"):
        return "meta"
    if col in (
        "Name Surname", "Contact Email", "Contact Phone Number",
        "Country", "Place of Residence (Municipality / City)",
        "This will help us to understand where our clients are living – Address",
        "Gender", "Year of Birth", "Level of Education",
        "Do you identify as Roma?", "If NO Roma – Connection to Roma Community",
    ):
        return "personal"
    return "business"


GROUP_COLORS = {
    "meta":     ("1F4E79", "D6E4F0"),
    "personal": ("375623", "E2EFDA"),
    "business": ("7B3F00", "FCE4D6"),
    "redi":     ("4B0082", "EDE7F6"),
}

GROUP_LABELS = {
    "meta":     "Meta Information",
    "personal": "Personal Information",
    "business": "Business Information",
    "redi":     "REDI Staff",
}

# ---------------------------------------------------------------------------
# Dropdown helper – вредностите се пишуваат во скриен sheet
# ---------------------------------------------------------------------------

def _build_lists_sheet(wb: Workbook) -> dict[str, str]:
    """
    Пишува dropdown листи во скриен sheet 'Lists'.
    Враќа речник {col_name: formula_string} за DataValidation.
    """
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"

    formulas = {}
    col_idx = 1
    for col_name, options in DROPDOWNS.items():
        col_letter = get_column_letter(col_idx)
        for row_idx, val in enumerate(options, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        last_row = len(options)
        formulas[col_name] = f"Lists!${col_letter}$1:${col_letter}${last_row}"
        col_idx += 1

    return formulas


# ---------------------------------------------------------------------------
# Главна функција
# ---------------------------------------------------------------------------

def create_template(output_path: str, num_rows: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"

    dv_formulas = _build_lists_sheet(wb)

    thin  = Side(style="thin", color="BFBFBF")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # --- Row 1: group banners ---
    groups_seq = [_col_group(c) for c in COLUMNS]
    col_idx = 1
    while col_idx <= len(COLUMNS):
        group = groups_seq[col_idx - 1]
        span_start = col_idx
        while col_idx <= len(COLUMNS) and groups_seq[col_idx - 1] == group:
            col_idx += 1
        span_end = col_idx - 1
        hdr_color, _ = GROUP_COLORS[group]
        label = GROUP_LABELS[group]
        start_letter = get_column_letter(span_start)
        end_letter   = get_column_letter(span_end)
        ws.merge_cells(f"{start_letter}1:{end_letter}1")
        cell = ws.cell(row=1, column=span_start, value=label)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = center
        cell.border    = bdr
        for c in range(span_start + 1, span_end + 1):
            ws.cell(row=1, column=c).border = bdr

    ws.row_dimensions[1].height = 22

    # --- Row 2: column headers ---
    for i, col_name in enumerate(COLUMNS, start=1):
        group = _col_group(col_name)
        hdr_color, _ = GROUP_COLORS[group]
        cell = ws.cell(row=2, column=i, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = center
        cell.border    = bdr

    ws.row_dimensions[2].height = 48
    ws.freeze_panes = "A3"

    # --- Data rows ---
    for row_num in range(3, 3 + num_rows):
        _, row_color_default = GROUP_COLORS["personal"]
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            group = _col_group(col_name)
            _, row_color = GROUP_COLORS[group]
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font      = Font(size=10)
            cell.border    = bdr
            cell.alignment = wrap
            if row_num % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=row_color)
        ws.row_dimensions[row_num].height = 18

    # --- Data validations (dropdowns) ---
    last_data_row = 2 + num_rows
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        if col_name in dv_formulas:
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list",
                formula1=f"={dv_formulas[col_name]}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid value",
                error="Please select a value from the dropdown list.",
            )
            dv.sqref = f"{col_letter}3:{col_letter}{last_data_row}"
            ws.add_data_validation(dv)

    # --- Column widths ---
    col_widths = {
        "REDI Staff Member Name": 22,
        "Date of Mapping": 16,
        "Name Surname": 22,
        "Contact Email": 26,
        "Contact Phone Number": 20,
        "Country": 14,
        "Place of Residence (Municipality / City)": 28,
        "This will help us to understand where our clients are living – Address": 32,
        "Gender": 12,
        "Year of Birth": 14,
        "Level of Education": 20,
        "Do you identify as Roma?": 18,
        "If NO Roma – Connection to Roma Community": 30,
        "Business Name": 24,
        "Address of the Business": 26,
        "Short Description of the Business": 34,
        "Business Sector": 26,
        "Is the Business Registered?": 18,
        "Date of Business Registration": 20,
        "Number of Employees (formal and informal aggregate)": 20,
        "Was Business Supported in Phase I?": 18,
        "Year of Initial Support": 16,
        "Type of Service Needed": 32,
        "Interested in Business Club Membership?": 20,
        "Previously had a Business Loan?": 18,
        "Informed about Measures / Help for Business": 26,
        "[REDI Staff] Assessment of Client Needs": 34,
        "[REDI Staff] Action Recommended": 32,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    # --- Legend sheet ---
    _add_legend_sheet(wb)

    wb.save(output_path)
    print(f"Template saved: {output_path}  ({num_rows} empty rows, {len(COLUMNS)} columns)")


def _add_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Legend")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45

    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    entries = [
        ("Color Group",                       "Meaning"),
        ("Dark Blue  / Light Blue",           "Meta Information (REDI Staff, Date)"),
        ("Dark Green / Light Green",          "Personal Information of the Client"),
        ("Dark Brown / Light Peach",          "Business Information"),
        ("Purple     / Light Purple",         "REDI Staff Assessment & Recommendation"),
    ]
    hdr_fills = [
        ("1F4E79", "D6E4F0"),
        ("375623", "E2EFDA"),
        ("7B3F00", "FCE4D6"),
        ("4B0082", "EDE7F6"),
    ]

    # Header row
    for col_idx, val in enumerate(entries[0], start=1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="333333")
        cell.border    = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for i, (label, desc) in enumerate(entries[1:], start=2):
        hdr_c, row_c = hdr_fills[i - 2]
        ca = ws.cell(row=i, column=1, value=label)
        ca.fill      = PatternFill("solid", fgColor=hdr_c)
        ca.font      = Font(color="FFFFFF", bold=True, size=10)
        ca.border    = bdr
        ca.alignment = Alignment(horizontal="left", vertical="center")

        cb = ws.cell(row=i, column=2, value=desc)
        cb.fill      = PatternFill("solid", fgColor=row_c)
        cb.font      = Font(size=10)
        cb.border    = bdr
        cb.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[i].height = 20


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--output", default="redi_questionnaire_template.xlsx",
                        help="Output Excel file (default: redi_questionnaire_template.xlsx)")
    parser.add_argument("--rows", type=int, default=100,
                        help="Number of empty data rows to pre-fill (default: 100)")
    args = parser.parse_args()
    create_template(args.output, args.rows)
