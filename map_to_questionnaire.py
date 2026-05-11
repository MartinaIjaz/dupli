"""
Ги пренесува податоците од постоечки Excel фајл во структурата на
"Draft 1 Mapping Questionnaire - Entrepreneurs" (REDI NGO).

Употреба:
    python3 map_to_questionnaire.py --input vashiot_fajl.xlsx --output questionnaire.xlsx

Скриптата автоматски ги препознава колоните и ги пренесува во прашалникот.
Полиња за кои нема податоци остануваат празни — за REDI staff да ги пополни.
"""

import argparse
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Точни имиња на колони од прашалникот
# ---------------------------------------------------------------------------

QUESTIONNAIRE_COLUMNS = [
    "REDI staff member name",
    "Date of mapping (today)",
    "Name Surname",
    "Contact Email",
    "Contact Phone number",
    "Country",
    "Place of Residence (Municipality, City)",
    "Address",
    "Gender",
    "Year of your birth",
    "What is your level of education?",
    "Do you identify as Roma?",
    "If NO, what is the connection to the Roma community?",
    "Business Name?",
    "Address of the Business",
    "Short description of the business. What is your business about?",
    "Which business sector does the business belong too?",
    "Is the business registered?",
    "Date of Business Registration?",
    "Number of Employees (formal and informal aggregate)",
    "Was your business supported in the Phase I?",
    "Year of Initial Support",
    "Which type of service do you need? (question for entrepreneurs)",
    "Are you interested in a Business Club Membership?",
    "Did you previously have a business loan?",
    "Informed about measure, or anything that helps his business",
    "[For REDI Staff] Describe your assessment of the clients needs",
    "[For REDI Staff] Action recommended?",
]

# ---------------------------------------------------------------------------
# Dropdown вредности за полиња со понудени одговори
# ---------------------------------------------------------------------------

DROPDOWN_OPTIONS: dict[str, list[str]] = {
    "Gender": ["Male", "Female", "Other"],
    "What is your level of education?": [
        "No formal education", "Elementary School", "High School",
        "University Student", "Bachelor", "Master", "PhD", "Vocational Training",
    ],
    "Do you identify as Roma?": ["Yes", "No"],
    "Which business sector does the business belong too?": [
        "Agriculture and animal husbandry", "Construction", "Services (Accounting, Law)",
        "Tourism and Hospitality", "Beauty and Health", "Transportation",
        "Production", "Trade", "Craftsmanship", "Recycling & Upcycling",
        "Cleaning Agency", "Other",
    ],
    "Is the business registered?": ["Yes", "No"],
    "Was your business supported in the Phase I?": ["Yes", "No"],
    "Year of Initial Support": ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027"],
    "Which type of service do you need? (question for entrepreneurs)": [
        "Incubation Program", "Acceleration Program", "Growth Program",
        "Digitalization of the business", "One-on-one Mentorship",
        "Business Registration", "Brand Identity", "Business Plan Writing",
        "Marketing and Promotion", "Networking",
        "Access to Grant from National Employment Service Agency",
        "Access to Grant from REDI", "Access to grant from IPARD",
        "Business Loan", "Certified Training", "Consulting",
        "Legal Support", "Accountancy Support",
        "Assistance at Finding Additional Workers",
        "Assistance at Starting a Business",
        "Internship Program Development",
        "Improved Business Practices",
        "Assistance in Accessing Government Institutions",
        "Assistance in Forming Cooperatives",
        "Social Impact and Corporate Social Responsibility (ESG)",
        "Nothing from above",
    ],
    "Are you interested in a Business Club Membership?": ["Yes", "No", "Maybe"],
    "Did you previously have a business loan?": ["Yes", "No"],
    "Informed about measure, or anything that helps his business": ["Yes", "No"],
    "[For REDI Staff] Action recommended?": [
        "Develop a business plan",
        "Link to a financial institution (MFI or Bank)",
        "Apply for REDI Grant Scheme Support",
        "Link to the Agency of Employment to find or subsidized employment",
        "Link to Government measures, to be self-employed",
        "Train on Business Growth and Development",
        "Business Formalization (Registration)",
        "One-on-One Expert Support",
        "Digitalization (Roma Digital Boost)",
        "Marketing and Promotion",
        "Networking (Business Clubs)",
        "Linking to potential employees",
        "Business Consultancy",
        "Link clients to training programs (certified or other)",
        "No Action for the moment",
        "Other",
    ],
}

# ---------------------------------------------------------------------------
# Авто-детекција: мап од можни имиња на влезни колони → поле на прашалник
# ---------------------------------------------------------------------------

COLUMN_MAP: dict[str, str] = {
    # Комбинирано ime + prezime
    "ime i prezime":            "Name Surname",
    "name surname":             "Name Surname",
    "full name":                "Name Surname",
    "полно ime":                "Name Surname",
    # Ime / Prezime одделно — се спојуваат
    "ime":                      "__ime__",
    "prezime":                  "__prezime__",
    "first name":               "__ime__",
    "last name":                "__prezime__",
    "name":                     "__ime__",
    "surname":                  "__prezime__",
    # Е-пошта
    "е-пошта":                  "Contact Email",
    "email":                    "Contact Email",
    "e-mail":                   "Contact Email",
    "e mail":                   "Contact Email",
    # Телефон
    "телефон":                  "Contact Phone number",
    "phone":                    "Contact Phone number",
    "phone number":             "Contact Phone number",
    "mobile":                   "Contact Phone number",
    # Земја
    "земја":                    "Country",
    "country":                  "Country",
    "држава":                   "Country",
    # Општина / Место
    "општина":                  "Place of Residence (Municipality, City)",
    "municipality":             "Place of Residence (Municipality, City)",
    "место":                    "Place of Residence (Municipality, City)",
    "град":                     "Place of Residence (Municipality, City)",
    "city":                     "Place of Residence (Municipality, City)",
    "place of residence":       "Place of Residence (Municipality, City)",
    # Адреса (лична)
    "адреса":                   "Address",
    "address":                  "Address",
    "улица":                    "Address",
    # Пол
    "пол":                      "Gender",
    "gender":                   "Gender",
    "sex":                      "Gender",
    # Година на раѓање
    "година на раѓање":         "Year of your birth",
    "year of birth":            "Year of your birth",
    "роденден":                 "Year of your birth",
    "birthdate":                "Year of your birth",
    "date of birth":            "Year of your birth",
    # Образование
    "образование":              "What is your level of education?",
    "level of education":       "What is your level of education?",
    "education":                "What is your level of education?",
    # Рома
    "рома":                     "Do you identify as Roma?",
    "roma":                     "Do you identify as Roma?",
    "is roma":                  "Do you identify as Roma?",
    # Врска со Рома
    "врска со ромска заедница": "If NO, what is the connection to the Roma community?",
    "connection to roma":       "If NO, what is the connection to the Roma community?",
    # Бизнис назив
    "бизнис":                   "Business Name?",
    "business name":            "Business Name?",
    "назив на фирма":           "Business Name?",
    "фирма":                    "Business Name?",
    # Адреса на бизнис
    "адреса на бизнис":         "Address of the Business",
    "business address":         "Address of the Business",
    "address of the business":  "Address of the Business",
    # Опис
    "опис на бизнис":           "Short description of the business. What is your business about?",
    "business description":     "Short description of the business. What is your business about?",
    "description":              "Short description of the business. What is your business about?",
    # Сектор
    "сектор":                   "Which business sector does the business belong too?",
    "sector":                   "Which business sector does the business belong too?",
    "business sector":          "Which business sector does the business belong too?",
    # Регистрација
    "регистриран":              "Is the business registered?",
    "registered":               "Is the business registered?",
    # Датум на регистрација
    "датум на регистрација":    "Date of Business Registration?",
    "date of registration":     "Date of Business Registration?",
    "registration date":        "Date of Business Registration?",
    # Вработени
    "вработени":                "Number of Employees (formal and informal aggregate)",
    "employees":                "Number of Employees (formal and informal aggregate)",
    "number of employees":      "Number of Employees (formal and informal aggregate)",
    # Фаза I
    "фаза i":                   "Was your business supported in the Phase I?",
    "phase i":                  "Was your business supported in the Phase I?",
    "phase 1":                  "Was your business supported in the Phase I?",
    # Година на поддршка
    "година на поддршка":       "Year of Initial Support",
    "year of support":          "Year of Initial Support",
    # Тип на услуга
    "тип на услуга":            "Which type of service do you need? (question for entrepreneurs)",
    "service needed":           "Which type of service do you need? (question for entrepreneurs)",
    "service":                  "Which type of service do you need? (question for entrepreneurs)",
    # Бизнис клуб
    "бизнис клуб":              "Are you interested in a Business Club Membership?",
    "business club":            "Are you interested in a Business Club Membership?",
    # Кредит
    "кредит":                   "Did you previously have a business loan?",
    "loan":                     "Did you previously have a business loan?",
    "business loan":            "Did you previously have a business loan?",
    # Информиран
    "информиран":               "Informed about measure, or anything that helps his business",
    "informed":                 "Informed about measure, or anything that helps his business",
    # REDI оценка
    "процена":                  "[For REDI Staff] Describe your assessment of the clients needs",
    "assessment":               "[For REDI Staff] Describe your assessment of the clients needs",
    # Препорака
    "потребни акции":           "[For REDI Staff] Action recommended?",
    "action":                   "[For REDI Staff] Action recommended?",
    "action recommended":       "[For REDI Staff] Action recommended?",
    # Датум на мапирање
    "датум":                    "Date of mapping (today)",
    "date":                     "Date of mapping (today)",
    "date of mapping":          "Date of mapping (today)",
}


def detect_mapping(df: pd.DataFrame):
    mapping = {}
    lower_cols = {c.lower().strip(): c for c in df.columns}
    ime_col = prez_col = None

    for lower, original in lower_cols.items():
        target = COLUMN_MAP.get(lower)
        if target == "__ime__":
            ime_col = original
        elif target == "__prezime__":
            prez_col = original
        elif target:
            mapping[original] = target

    return mapping, ime_col, prez_col


def _first_non_empty(series: pd.Series):
    for val in series:
        if pd.notna(val) and str(val).strip() not in ("", "nan", "None"):
            return val
    return None


def build_questionnaire(df, mapping, ime_col, prez_col) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        qrow = {col: None for col in QUESTIONNAIRE_COLUMNS}

        if ime_col and prez_col:
            ime  = str(row[ime_col]).strip()  if pd.notna(row[ime_col])  else ""
            prez = str(row[prez_col]).strip() if pd.notna(row[prez_col]) else ""
            full = f"{ime} {prez}".strip()
            if full:
                qrow["Name Surname"] = full
        elif ime_col and pd.notna(row[ime_col]):
            qrow["Name Surname"] = row[ime_col]

        for src_col, dst_col in mapping.items():
            val = row[src_col]
            if pd.notna(val) and str(val).strip() not in ("", "nan", "None"):
                if qrow[dst_col] is None:
                    qrow[dst_col] = val

        rows.append(qrow)

    return pd.DataFrame(rows, columns=QUESTIONNAIRE_COLUMNS)


# ---------------------------------------------------------------------------
# Боени групи на колони
# ---------------------------------------------------------------------------

def _col_group(col: str) -> str:
    if col.startswith("[For REDI"):
        return "redi"
    if col in ("REDI staff member name", "Date of mapping (today)"):
        return "meta"
    if col in (
        "Name Surname", "Contact Email", "Contact Phone number", "Country",
        "Place of Residence (Municipality, City)", "Address", "Gender",
        "Year of your birth", "What is your level of education?",
        "Do you identify as Roma?",
        "If NO, what is the connection to the Roma community?",
    ):
        return "personal"
    return "business"


GROUP_COLORS = {
    "meta":     ("1F4E79", "D6E4F0"),
    "personal": ("375623", "E2EFDA"),
    "business": ("7B3F00", "FCE4D6"),
    "redi":     ("4B0082", "EDE7F6"),
}


# ---------------------------------------------------------------------------
# Excel форматирање + dropdown валидација
# ---------------------------------------------------------------------------

def _add_dropdowns(ws, df: pd.DataFrame, max_rows: int = 500) -> None:
    """Додај dropdown листи за прашања со понудени одговори."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        options = DROPDOWN_OPTIONS.get(col_name)
        if not options:
            continue

        # За кратки листи — inline formula
        quoted = [f'"{o}"' for o in options]
        formula = ",".join(quoted)
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{max_rows}"

        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"' if len(formula) <= 255 else None,
            allow_blank=True,
            showErrorMessage=False,
        )
        # Ако е предолга листата, скрати ги опциите во формулата
        short_options = options[:10] if len(formula) > 255 else options
        dv.formula1 = '"' + ",".join(short_options) + '"'
        dv.sqref = cell_range
        ws.add_data_validation(dv)


def format_sheet(ws, df: pd.DataFrame) -> None:
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Заглавија
    for col_idx, col_name in enumerate(df.columns, start=1):
        group = _col_group(col_name)
        hdr_color, _ = GROUP_COLORS[group]
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = col_name
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = center
        cell.border    = bdr

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 50

    # Редови со податоци
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(df.columns, start=1):
            group = _col_group(col_name)
            _, row_color = GROUP_COLORS[group]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font      = Font(size=10)
            cell.border    = bdr
            cell.alignment = wrap
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=row_color)
        ws.row_dimensions[row_idx].height = 18

    # Ширина на колони
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(col_name),
            max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)),
                default=0,
            ),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 38)

    _add_dropdowns(ws, df)


def add_legend_sheet(wb) -> None:
    ws = wb.create_sheet("Legend")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 45

    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    header_data = [("Color Group", "Fields")]
    rows_data = [
        ("1F4E79", "D6E4F0", "Dark Blue / Light Blue",    "REDI Staff Name, Date of Mapping"),
        ("375623", "E2EFDA", "Dark Green / Light Green",  "Personal data: Name, Email, Phone, Country, Address, Gender, Education, Roma..."),
        ("7B3F00", "FCE4D6", "Dark Brown / Light Peach",  "Business data: Business Name, Sector, Registration, Employees, Services..."),
        ("4B0082", "EDE7F6", "Purple / Light Purple",     "[For REDI Staff] Assessment & Action Recommended"),
    ]

    for col_idx, val in enumerate(header_data[0], start=1):
        c = ws.cell(row=1, column=col_idx, value=val)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="333333")
        c.border = bdr
        c.alignment = center

    for i, (hc, rc, label, desc) in enumerate(rows_data, start=2):
        ca = ws.cell(row=i, column=1, value=label)
        ca.fill = PatternFill("solid", fgColor=hc)
        ca.font = Font(color="FFFFFF", bold=True)
        ca.border = bdr
        ca.alignment = left

        cb = ws.cell(row=i, column=2, value=desc)
        cb.fill = PatternFill("solid", fgColor=rc)
        cb.border = bdr
        cb.alignment = left
        ws.row_dimensions[i].height = 22


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input",  required=True)
    parser.add_argument("--output", default="questionnaire_output.xlsx")
    parser.add_argument("--sheet",  default=0)
    args = parser.parse_args()

    try:
        sheet = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
        df = pd.read_excel(args.input, sheet_name=sheet)
    except FileNotFoundError:
        sys.exit(f"Error: file '{args.input}' not found.")
    except Exception as exc:
        sys.exit(f"Error reading file: {exc}")

    print(f"Loaded {len(df)} rows.")
    print(f"Input columns: {list(df.columns)}")

    mapping, ime_col, prez_col = detect_mapping(df)

    print("\nColumn mappings detected:")
    if ime_col and prez_col:
        print(f"  '{ime_col}' + '{prez_col}'  →  Name Surname")
    elif ime_col:
        print(f"  '{ime_col}'  →  Name Surname")
    for src, dst in mapping.items():
        print(f"  '{src}'  →  '{dst}'")

    df_q = build_questionnaire(df, mapping, ime_col, prez_col)

    filled = df_q.notna().sum().sum()
    total  = len(df_q) * len(QUESTIONNAIRE_COLUMNS)
    pct    = 100 * filled // total if total else 0
    print(f"\nFilled cells: {filled} / {total} ({pct}%)")
    empty_cols = [c for c in QUESTIONNAIRE_COLUMNS if df_q[c].isna().all()]
    if empty_cols:
        print(f"Empty columns (no source data): {len(empty_cols)} columns")

    try:
        with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
            df_q.to_excel(writer, index=False, sheet_name="Questionnaire")
            format_sheet(writer.sheets["Questionnaire"], df_q)

        wb = load_workbook(args.output)
        add_legend_sheet(wb)
        wb.save(args.output)

        print(f"\nSaved: {args.output}")
    except Exception as exc:
        sys.exit(f"Error saving file: {exc}")


if __name__ == "__main__":
    main()
